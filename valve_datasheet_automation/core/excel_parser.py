from pathlib import Path
from typing import Dict, List, Any
import openpyxl

class ExcelParser:
    """
    Parses Excel files to extract structured data.
    """
    def parse_field_applicability(self, file_path: Path, sheet_name: str = "Field Applicability") -> Dict[str, List[str]]:
        """
        Parses an Excel sheet to determine which fields are applicable for different VDS templates.

        Assumes the Excel sheet has the following structure:
        - Row 1 contains template identifiers (e.g., "Ball Valve", "Gate Valve")
        - Column 1 contains field names (e.g., "piping_class", "body_material")
        - Cells indicate applicability (e.g., "Y" for applicable)

        Example sheet:
        Template/Field | Ball Valve | Gate Valve | Check Valve
        ------------------------------------------------------
        piping_class   | Y          | Y          | Y
        body_material  | Y          | Y          | Y
        ball_material  | Y          | N          | N
        disc_material  | N          | N          | Y

        Args:
            file_path (Path): Path to the Excel file.
            sheet_name (str): Name of the sheet to parse.

        Returns:
            Dict[str, List[str]]: A dictionary where keys are template identifiers
                                  and values are lists of applicable field names.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")

        sheet = workbook[sheet_name]
        
        # Read header row for template identifiers
        template_identifiers = []
        for cell in sheet[1]:
            if cell.value and cell.column > 1: # Skip first column (field names)
                template_identifiers.append(str(cell.value).strip())
        
        if not template_identifiers:
            raise ValueError("No template identifiers found in the first row of the sheet.")

        # Read first column for field names
        field_names = []
        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value:
                field_names.append(str(cell_value).strip())
        
        if not field_names:
            raise ValueError("No field names found in the first column of the sheet.")

        field_applicability: Dict[str, List[str]] = {template: [] for template in template_identifiers}

        # Populate applicability
        for row_idx, field_name in enumerate(field_names, start=2): # Start from row 2 (after header)
            for col_idx, template_id in enumerate(template_identifiers, start=2): # Start from col 2 (after field names)
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if str(cell_value).strip().upper() == "Y":
                    field_applicability[template_id].append(field_name)

        return field_applicability

