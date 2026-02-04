"""
Extract all valve datasheet Excel files to JSON format.
This enables 100% accurate field mapping for all valve types.
"""

import json
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional


def extract_vds_sheet_data(ws, vds_no: str) -> Dict[str, Any]:
    """Extract all field data from a single VDS sheet."""
    data = {
        "vds_no": vds_no,
        "fields": {}
    }

    current_section = None

    for row in ws.iter_rows(min_row=1, max_row=60, max_col=6, values_only=True):
        # Get cell values
        col_a = str(row[0]).strip() if row[0] else ""
        col_b = str(row[1]).strip() if row[1] else ""
        col_c = str(row[2]).strip() if row[2] else ""

        # Skip empty rows
        if not col_a and not col_b and not col_c:
            continue

        # Detect field patterns
        # Pattern 1: Field in column A, Value in column C
        if col_a and col_c and col_a not in ['#VALUE!', 'Construction', 'Material', 'Material ']:
            field_name = col_a.replace(':', '').strip()
            data["fields"][field_name] = col_c

        # Pattern 2: Section header in A, Sub-field in B, Value in C
        elif col_a in ['Construction', 'Material', 'Material ']:
            current_section = col_a.strip()
            if col_b and col_c:
                field_name = f"{current_section} - {col_b}".replace(':', '').strip()
                data["fields"][field_name] = col_c

        # Pattern 3: Sub-field continuation (A empty, B has field, C has value)
        elif not col_a and col_b and col_c:
            if current_section:
                field_name = f"{current_section} - {col_b}".replace(':', '').strip()
            else:
                field_name = col_b.replace(':', '').strip()
            data["fields"][field_name] = col_c

    return data


def extract_index_sheet(ws) -> List[Dict[str, Any]]:
    """Extract Index sheet data."""
    rows_data = []
    headers = []

    for i, row in enumerate(ws.iter_rows(max_row=200, max_col=12, values_only=True)):
        if i == 0:
            # First row is headers
            headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
            continue

        # Skip title rows
        if row[1] in [None, 'BALL VALVE', 'GATE VALVE', 'GLOBE VALVE', 'CHECK VALVE',
                      'BUTTERFLY VALVE', 'DBB VALVE', 'Needle Valve', 'Sheet Index']:
            continue

        # Skip if no VDS number
        if not row[1]:
            continue

        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers) and val is not None:
                row_dict[headers[j]] = str(val).strip() if val else ""

        if row_dict.get('VDS'):
            rows_data.append(row_dict)

    return rows_data


def extract_excel_file(excel_path: Path) -> Dict[str, Any]:
    """Extract all data from an Excel file."""
    print(f"Extracting: {excel_path.name}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)

    result = {
        "file": excel_path.name,
        "valve_type": excel_path.stem.split('-')[0],
        "index": [],
        "datasheets": {}
    }

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('~$'):
            continue

        ws = wb[sheet_name]

        if sheet_name.lower() == 'index':
            result["index"] = extract_index_sheet(ws)
            print(f"  - Index: {len(result['index'])} entries")
        else:
            # Individual VDS sheet
            vds_data = extract_vds_sheet_data(ws, sheet_name)
            if vds_data["fields"]:
                result["datasheets"][sheet_name] = vds_data

    print(f"  - Datasheets: {len(result['datasheets'])} VDS sheets")
    wb.close()

    return result


def normalize_field_name(field: str) -> str:
    """Normalize field names for consistent mapping."""
    # Clean up the field name first
    field = field.lower().strip()
    # Replace em-dash and en-dash with regular hyphen
    field = field.replace('–', '-').replace('—', '-')
    # Remove extra spaces
    field = ' '.join(field.split())

    # Map common variations to standard field names
    mappings = {
        'vds no': 'vds_no',
        'piping class': 'piping_class',
        'size range': 'size_range',
        'valve type': 'valve_type',
        'valve standard': 'valve_standard',
        'pressure class': 'pressure_class',
        'design pressure': 'design_pressure',
        'corrosion allowance': 'corrosion_allowance',
        'sour service requirements': 'sour_service',
        'sour service': 'sour_service',
        'end connections': 'end_connections',
        'face to face dimension': 'face_to_face',
        'face to face': 'face_to_face',
        'operation': 'operation',
        'service': 'service',
        # Construction fields
        'construction - body': 'body_construction',
        'construction - ball': 'ball_construction',
        'construction - wedge': 'wedge_construction',
        'construction - disc': 'disc_construction',
        'construction - stem': 'stem_construction',
        'construction - seat': 'seat_construction',
        'construction - back seat': 'back_seat_construction',
        'construction - packing': 'packing_construction',
        'construction - shaft': 'shaft_construction',
        'construction - locks': 'locks',
        # Material fields - handle various formats
        'material - body material': 'body_material',
        'material  - body material': 'body_material',
        'material - body': 'body_material',
        'material - ball material': 'ball_material',
        'material - ball': 'ball_material',
        'material - wedge material': 'wedge_material',
        'material - wedge': 'wedge_material',
        'material - disc material': 'disc_material',
        'material - disc': 'disc_material',
        'material - trim material': 'trim_material',
        'material - trim': 'trim_material',
        'material - seat material': 'seat_material',
        'material - seat': 'seat_material',
        'material - seal material': 'seal_material',
        'material - seal': 'seal_material',
        'material - back seat mat.': 'back_seat_material',
        'material - back seat material': 'back_seat_material',
        'material - back seat': 'back_seat_material',
        'material - stem material': 'stem_material',
        'material - stem': 'stem_material',
        'material - gland material': 'gland_material',
        'material - gland': 'gland_material',
        'material - gland packing': 'gland_packing',
        'material - packing': 'gland_packing',
        'material - lever / handwheel': 'lever_handwheel',
        'material - lever/handwheel': 'lever_handwheel',
        'material - handwheel': 'lever_handwheel',
        'material - lever': 'lever_handwheel',
        'material - spring': 'spring_material',
        'material - spring material': 'spring_material',
        'material - spring': 'spring_material',
        'material - shaft material': 'shaft_material',
        'material - shaft': 'shaft_material',
        'material - gaskets': 'gaskets',
        'material - gasket': 'gaskets',
        'material - bolts': 'bolts',
        'material - bolt': 'bolts',
        'material - nuts': 'nuts',
        'material - nut': 'nuts',
        # Testing fields - handle various formats including em-dash versions
        'marking - purchaser\'s specification': 'marking_purchaser',
        'marking - purchaser': 'marking_purchaser',
        'marking - manufacturer': 'marking_manufacturer',
        'inspection - testing': 'inspection_testing',
        'inspection': 'inspection_testing',
        'leakage rate': 'leakage_rate',
        # Test pressure fields - handle full names
        'test pressure - hydrostatic shell': 'hydrotest_shell',
        'test pressure - hydrostatic closure': 'hydrotest_closure',
        'test pressure - pneumatic lp': 'pneumatic_test',
        'hydrotest shell test pressure': 'hydrotest_shell',
        'hydrotest closure test pressure': 'hydrotest_closure',
        'pneumatic lp test pressure': 'pneumatic_test',
        'hydrostatic shell': 'hydrotest_shell',
        'hydrostatic closure': 'hydrotest_closure',
        'pneumatic lp': 'pneumatic_test',
        'material certification': 'material_certification',
        'fire rating': 'fire_rating',
        'finish / general specification': 'finish',
        'finish': 'finish',
    }

    # Check direct match first
    if field in mappings:
        return mappings[field]

    # Try to match after normalizing underscores and hyphens
    normalized = field.replace(' ', '_').replace('-', '_')
    # Remove multiple consecutive underscores
    while '__' in normalized:
        normalized = normalized.replace('__', '_')
    normalized = normalized.strip('_')

    # Check if the normalized version matches any known patterns
    additional_mappings = {
        'material___seal_material': 'seal_material',
        'marking___purchaser\'s_specification': 'marking_purchaser',
        'marking_purchaser\'s_specification': 'marking_purchaser',
        'marking__manufacturer': 'marking_manufacturer',
        'inspection__testing': 'inspection_testing',
        'hydrotest_shell_test_pressure': 'hydrotest_shell',
        'hydrotest_closure_test_pressure': 'hydrotest_closure',
        'pneumatic_lp_test_pressure': 'pneumatic_test',
        # Shaft field variations
        'construction_shaft': 'shaft_construction',
        'material_shaft_material': 'shaft_material',
        'material_shaft': 'shaft_material',
        # Disc field variations
        'material_disc_material': 'disc_material',
        'material_disc': 'disc_material',
        'construction_disc': 'disc_construction',
        # Wedge field variations
        'material_wedge_material': 'wedge_material',
        'material_wedge': 'wedge_material',
        'construction_wedge': 'wedge_construction',
        # Trim field variations
        'material_trim_material': 'trim_material',
        'material_trim': 'trim_material',
    }

    if normalized in additional_mappings:
        return additional_mappings[normalized]

    # Remove apostrophes and clean up any remaining special characters
    normalized = normalized.replace("'", "").replace("'", "")

    # Final check for purchaser specification variations
    if 'purchaser' in normalized and 'specification' in normalized:
        return 'marking_purchaser'

    return normalized


def create_unified_vds_index(all_data: List[Dict]) -> Dict[str, Dict]:
    """Create a unified VDS index from all extracted data."""
    unified = {}

    for file_data in all_data:
        # Add from individual datasheets (most detailed)
        for vds_no, vds_data in file_data.get("datasheets", {}).items():
            normalized_fields = {}
            for field, value in vds_data.get("fields", {}).items():
                norm_field = normalize_field_name(field)
                normalized_fields[norm_field] = value

            unified[vds_no] = {
                "vds_no": vds_no,
                "source_file": file_data["file"],
                "valve_type": file_data["valve_type"],
                **normalized_fields
            }

        # Also add from index (backup for any missing)
        for idx_row in file_data.get("index", []):
            vds_no = idx_row.get("VDS", "")
            if vds_no and vds_no not in unified:
                unified[vds_no] = {
                    "vds_no": vds_no,
                    "source_file": file_data["file"],
                    "valve_type": file_data["valve_type"],
                    "size_range": idx_row.get("Size Range", ""),
                    "valve_type_desc": idx_row.get("Valve type", ""),
                    "end_connections": idx_row.get("End Connections", ""),
                }

    return unified


def main():
    """Main extraction function."""
    valve_data_dir = Path("unstructured/VALVE DATA SHEET")
    output_dir = Path("unstructured")

    if not valve_data_dir.exists():
        print(f"Directory not found: {valve_data_dir}")
        return

    all_data = []

    # Extract all Excel files
    excel_files = list(valve_data_dir.glob("*.xlsm"))
    print(f"Found {len(excel_files)} Excel files to extract\n")

    for excel_file in excel_files:
        if excel_file.name.startswith('~$'):
            continue

        try:
            file_data = extract_excel_file(excel_file)
            all_data.append(file_data)

            # Save individual file JSON
            output_file = output_dir / f"{excel_file.stem}_extracted.json"
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(file_data, f, indent=2, ensure_ascii=False)
            print(f"  -> Saved: {output_file.name}\n")

        except Exception as e:
            print(f"  ERROR: {e}\n")

    # Create unified VDS index
    print("Creating unified VDS index...")
    unified_index = create_unified_vds_index(all_data)

    unified_file = output_dir / "all_valve_vds_index.json"
    with open(unified_file, 'w', encoding='utf-8') as f:
        json.dump(unified_index, f, indent=2, ensure_ascii=False)

    print(f"\nUnified index saved: {unified_file}")
    print(f"Total VDS entries: {len(unified_index)}")

    # Print summary
    print("\n=== Extraction Summary ===")
    for file_data in all_data:
        print(f"{file_data['valve_type']}: {len(file_data['datasheets'])} datasheets")


if __name__ == "__main__":
    main()
